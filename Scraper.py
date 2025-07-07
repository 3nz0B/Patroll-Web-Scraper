import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
import json

def main_scraper():
    # Set up headless Chrome for main navigation
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--window-size=1920,1080")
    driver = webdriver.Chrome(options=options)

    # Set up separate scraper driver for contest detail pages
    scraper = webdriver.Chrome(options=options)

    # Target URL
    url = "https://patroll.unifiedpatents.com/contests?category=won"
    driver.get(url)

    # Data containers
    results = []
    max_pages = 19
    prefix = 'https://www.google.com'

    try:
        for page_num in range(1, max_pages + 1):
            print(f"🔄 Processing page {page_num}...")

            time.sleep(1)
            soup = BeautifulSoup(driver.page_source, "html.parser")
            ul = soup.find("ul", class_="ant-list-items")

            if not ul:
                break

            # Extract links
            temp_links = [a['href'] for a in ul.find_all('a', href=True)]
            contest_links = ["https://patroll.unifiedpatents.com" + link for link in temp_links if link.startswith('/contests/')]
            troll_patents = [link.split('/')[-1] for link in temp_links if link.startswith(prefix)]

            for idx, contest_url in enumerate(contest_links):
                print(f"🔍 Contest #{idx+1}: {contest_url}")
                try:
                    title = scrape_contest_title(contest_url, scraper)
                except:
                    title = "N/A"

                try:
                    prior_arts = scrape_prior_art(contest_url, scraper)
                except:
                    prior_arts = []

                parsed_prior_art = []
                
                try :
                    for art in prior_arts:
                        parsed_prior_art.append({
                            "patent_id": art,
                            "country_code": art[:2]  # US, EP, WO, etc.
                        })
                except :
                    print ("Error")

                results.append({
                    "contest_title": title,
                    "troll_patent_id": troll_patents[idx] if idx < len(troll_patents) else "N/A",
                    "prior_art_patents": parsed_prior_art,
                    "contest_url": contest_url
                })

            # Click "Next Page"
            try:
                next_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "li.ant-pagination-next[title='Next Page']"))
                )
                driver.execute_script("arguments[0].scrollIntoView();", next_button)
                driver.execute_script("arguments[0].click();", next_button)
            except Exception as e:
                print(f"⚠️ Could not find/click 'Next Page' button: {e}")
                break

    finally:
        driver.quit()
        scraper.quit()

        # Write JSON
        with open("scraped_patents.json", "w") as f:
            json.dump(results, f, indent=2)
        print("\n✅ Data saved to 'scraped_patents.json'")

def scrape_contest_title(contestlink,driver):
    
    #use contest link to get the title of the contest
    #Here is an example of a link you would recieve as a function parameter: https://patroll.unifiedpatents.com/contests/mJ5QT4wkDCCjhy9xb

    driver.get(contestlink)
    time.sleep(1)
    soup = BeautifulSoup(driver.page_source, "html.parser")

    #locate each contest
    h1 = soup.find("h1")
     
    if h1:
        title = h1.text.strip()
    else:
        title = 0
    print(title)
    return title

def scrape_prior_art_link(contestlink,driver):   
    try:
        driver.get(contestlink)
        

        
        try:
            download_link_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'DOWNLOAD WINNING PRIOR ART HERE:')]/following-sibling::a"))
            )
            download_link = download_link_element.get_attribute("href")
            return download_link
        except:
            print("Tag can't be found :(")
            return None

    except Exception as e:
        
        return None
    finally:
        print("Done")

def scrape_prior_art(contestlink,driver):
    
    #use contest link to get the ID of the prior art
    #Here is an example of a link you would recieve as a function parameter: https://patroll.unifiedpatents.com/contests/mJ5QT4wkDCCjhy9xb
   
    artlink=priorartlink(contestlink,driver)
    if artlink:
        pass
    else:
        return None

    prior_art_list = []

    if True:
        driver.get(artlink)
        
        #waits until page loads dynamically
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, 'body')) 
        
        )
        soup = BeautifulSoup(driver.page_source, "html.parser")

        p_tags = soup.find_all('p')

            

        for p_tag in p_tags:
                if "Winning Submissions:" in p_tag.text:
                    submissions_text = p_tag.text.split("Winning Submissions:")[1].strip()
                    #splits by semi colon to find list of prior art
                    references = [ref.strip() for ref in submissions_text.split(';')]
                    prior_art_list.extend(references)


    #this is another style of the page: https://www.unifiedpatents.com/insights/2025/3/31/3000-awarded-in-second-cloud-native-heroes-challenge-on-patroll
    if len(prior_art_list)==0:
        patent_links = soup.find('ul', {'data-rte-list': 'default'}).find_all('a')

        for link in patent_links:
            prior_art_list.append(link.text)

    print(prior_art_list)
    return '; '.join(prior_art_list)

def scrape_contests():

    # Load the Excel file
    excel_path = "output.xlsx"
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print("Error loading Excel file:", e)
        raise

    # Create a new sheet for the output if it doesn't exist
    sheet_name = "Scraped Contests"
    if sheet_name not in workbook.sheetnames:
        ws = workbook.create_sheet(sheet_name)
        ws.append(["Troll Patent", "Prior Art Patents", "Contest Title", "Award Amount", "Contest URL"])
    else:
        ws = workbook[sheet_name]

    # Setup Selenium browser
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--window-size=1920,1080")
    driver = webdriver.Chrome(options=options)

    base_url = "https://patroll.unifiedpatents.com"
    main_url = f"{base_url}/contests?category=finished"
    driver.get(main_url)

    max_pages = 2  # Adjust if you want more contests
    try:
        for page in range(1, max_pages + 1):
            print(f"Processing contest listings page {page}...")

            time.sleep(2)
            soup = BeautifulSoup(driver.page_source, "html.parser")
            contest_links = soup.select("ul.ant-list-items a[href^='/contests/']")
            contest_urls = list(set([base_url + a['href'] for a in contest_links]))

            for contest_url in contest_urls:
                print(f"Opening contest: {contest_url}")
                driver.get(contest_url)
                time.sleep(2)

                soup = BeautifulSoup(driver.page_source, "html.parser")

                # Contest Title
                title_elem = soup.find("h1")
                title = title_elem.text.strip() if title_elem else "N/A"

                # Troll Patent
                troll_patent = "N/A"
                troll_patent_link = soup.find("a", href=True)
                if troll_patent_link and "google.com" in troll_patent_link["href"]:
                    troll_patent = troll_patent_link.text.strip()

                # Award Amount
                award_elem = soup.find("div", text=lambda t: t and "Award Amount" in t)
                award_amount = award_elem.find_next("div").text.strip() if award_elem else "N/A"

                # Prior Art Patents
                prior_art = []
                all_links = soup.find_all("a", href=True)
                for a in all_links:
                    href = a["href"]
                    if href.startswith("https://www.google.com/patents"):
                        prior_art.append(href.split("patents/")[-1])

                prior_art_str = ", ".join(set(prior_art))

                # Add row to Excel
                ws.append([troll_patent, prior_art_str, title, award_amount, contest_url])

            # Next Page
            try:
                next_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f"//li[@title='{page + 1}']"))
                )
                driver.execute_script("arguments[0].scrollIntoView();", next_button)
                driver.execute_script("arguments[0].click();", next_button)
            except:
                print("No more pages.")
                break

    finally:
        driver.quit()
        workbook.save(excel_path)
        print(f"\n✅ Data saved directly to '{sheet_name}' in Excel file.")

main_scraper()
